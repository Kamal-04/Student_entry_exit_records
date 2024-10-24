# import cv2
# import numpy as np
# from pyzbar.pyzbar import decode
# import time
# from datetime import datetime
# import xlsxwriter as xw
# import pandas as pd
# from openpyxl import load_workbook as ld
# from playsound import playsound


# # Create a new Excel workbook to store entries and exits
# demo = xw.Workbook("Demo.xlsx")
# sheet1 = demo.add_worksheet("sheet_1")

# # Define column indices for writing entry and exit data
# entry_id = 0
# entry_time = 1
# exit_id = 2
# exit_time = 3

# # Initialize video capture from webcam
# cap = cv2.VideoCapture(0)
# cap.set(3, 640)  # Set frame width
# cap.set(4, 480)  # Set frame height

# # List to keep track of students who have entered
# li = ['0']
# i = 1
# row_1 = 1  # Row to write entry data
# row_2 = 1  # Row to write exit data

# # Set headers for the Excel sheet
# sheet1.write(0, 0, "ENTRY ID")
# sheet1.write(0, 1, "NAME")
# sheet1.write(0, 2, "ENTRY TIME")
# sheet1.write(0, 3, "EXIT ID")
# sheet1.write(0, 4, "NAME")
# sheet1.write(0, 5, "EXIT TIME")

# # Load student data from the Excel file "LIL.xls"

# data = pd.read_excel("class_data.xlsx",engine='openpyxl') # Use the correct file extension and engine


# # Main loop for video capture and barcode detection
# while True:
#     success, img = cap.read()  # Capture the image
#     cv2.imshow('Result', img)  # Display the image on the screen

#     for barcode in decode(img):  # Decode the barcode from the image
#         myData = barcode.data.decode('utf-8')  # Get the barcode data
#         now = datetime.now()
#         current_time = now.strftime("%H:%M:%S")  # Get current time

#         flag = 0  # To track if a valid student ID is found
#         myName = "UNKNOWN"

#         # Iterate over the student data to find the matching ID
#         for index, row in data.iterrows():
#             if row['STUDENT ID'] == myData:
#                 myName = row['NAME']
#                 flag = 1
#                 break

#         # Check if the student is entering or exiting
#         if myData not in li:  # New entry (student entering)
#             li.append(myData)
#             sheet1.write(row_1, 0, myData)
#             sheet1.write(row_1, 1, myName)
#             sheet1.write(row_1, 2, current_time)
#             row_1 += 1  # Move to the next row for the next entry

#             # Play appropriate sound
#             if flag == 1:
#                 playsound("Welcome.mp3")  # Play welcome sound for recognized student
#             else:
#                 playsound("Wrong Buzzer Sound - Reject.mp3")  # Play error sound for unrecognized student
            
#             time.sleep(2.5)  # Delay for sound to play and avoid multiple detections
#         else:  # Student exiting
#             sheet1.write(row_2, 3, myData)
#             sheet1.write(row_2, 4, myName)
#             sheet1.write(row_2, 5, current_time)
#             li.remove(myData)
#             playsound("Exit.mp3")  # Play exit sound
#             row_2 += 1  # Move to the next row for the next exit
#             time.sleep(2.5)  # Delay for sound to play

#     # Exit the loop when the Enter key (keycode 10) is pressed
#     key = cv2.waitKey(1)
#     if key == 10:
#         break

# # # Close the Excel workbook when done
# demo.close()

import cv2
import numpy as np
from pyzbar.pyzbar import decode
import time
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
from playsound import playsound

# Initialize video capture from webcam
cap = cv2.VideoCapture(0)
cap.set(3, 640)  # Set frame width
cap.set(4, 480)  # Set frame height

# List to keep track of students who have entered
li = ['0']
i = 1
entry_row = 1  # Row to write entry data
exit_row = 1  # Row to write exit data

# Load student data from the Excel file "class_data.xlsx"
data = pd.read_excel("class_data.xlsx", engine='openpyxl')  # Use the correct file extension and engine

# Load or create the workbook
file_name = "Demo.xlsx"
try:
    workbook = load_workbook(file_name)
    sheet = workbook.active
except FileNotFoundError:
    # Create a new Excel workbook if it doesn't exist
    workbook = xw.Workbook(file_name)
    sheet = workbook.add_worksheet("sheet_1")
    # Set headers for the Excel sheet
    sheet.write(0, 0, "ENTRY ID")
    sheet.write(0, 1, "NAME")
    sheet.write(0, 2, "ENTRY TIME")
    sheet.write(0, 3, "EXIT ID")
    sheet.write(0, 4, "NAME")
    sheet.write(0, 5, "EXIT TIME")

# Main loop for video capture and barcode detection
while True:
    success, img = cap.read()  # Capture the image
    cv2.imshow('Result', img)  # Display the image on the screen

    for barcode in decode(img):  # Decode the barcode from the image
        myData = barcode.data.decode('utf-8')  # Get the barcode data
        now = datetime.now()
        current_time = now.strftime("%H:%M:%S")  # Get current time

        flag = 0  # To track if a valid student ID is found
        myName = "UNKNOWN"

        # Iterate over the student data to find the matching ID
        for index, row in data.iterrows():
            if row['STUDENT ID'] == myData:
                myName = row['NAME']
                flag = 1
                break

        # Check if the student is entering or exiting
        if myData not in li:  # New entry (student entering)
            li.append(myData)
            
            # Append entry data to the Excel sheet
            sheet.append([myData, myName, current_time, '', '', ''])
            entry_row += 1  # Move to the next row for the next entry

            # Play appropriate sound
            if flag == 1:
                playsound("Welcome.mp3")  # Play welcome sound for recognized student
            else:
                playsound("Wrong Buzzer Sound - Reject.mp3")  # Play error sound for unrecognized student
            
            time.sleep(2.5)  # Delay for sound to play and avoid multiple detections
        else:  # Student exiting
            # Append exit data to the Excel sheet
            sheet.append(['', '', '', myData, myName, current_time])
            li.remove(myData)
            playsound("Exit.mp3")  # Play exit sound
            exit_row += 1  # Move to the next row for the next exit
            time.sleep(2.5)  # Delay for sound to play

        # Save the workbook after each write
        workbook.save(file_name)

    # Exit the loop when the Enter key (keycode 10) is pressed
    key = cv2.waitKey(1)
    if key == 10:
        break

# Release the video capture and close the workbook
cap.release()
cv2.destroyAllWindows()
workbook.save(file_name)  # Save the workbook one final time

